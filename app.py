from flask import Flask, render_template, request
from datetime import datetime
from openpyxl import load_workbook


app = Flask(__name__)
# Главная страница: 
@app.route("/")
@app.route("/<string:user_name>")
def index(user_name = ""):
    user_name = user_name.capitalize()
    note_file = open("notes.txt", "r", encoding="utf-8")
    lst_note = [line for line in note_file]
    note_file.close()
    return render_template("index.html", lst_note = lst_note, user_name = user_name)

#  Страница добавления ЗАМЕТОК: дата и заметка в xlsx и txt
@app.route("/form", methods=["POST", "GET"])
def form():
    if request.method == "POST":
        data = request.form.get("dt")
        note = request.form.get("new_note")
        # Запись в txt:
        notes_file = open("notes.txt", "a+", encoding="utf-8")
        notes_file.write(str(data) + " " + str(note) + "\n")
        notes_file.close()
        # Запись в xlsx - Лист Заметки:
        excel = load_workbook("mybase.xlsx")
        page = excel["Заметки"]
        page.append([data, note])
        excel.save("mybase.xlsx")
        return render_template("susses.html")

    # Вариант отображения txt файла (!ИСПРАВЛЕНО: encoding='utf-8-sig')
    notes_file = open("notes.txt", "r+", encoding='utf-8-sig')
    rows = [[datetime.strptime(row[:10], "%Y-%m-%d"), row[10:].strip()] for row in notes_file]
    notes_file.close()
    return render_template("form.html", rows=rows)

#  Страница добавления ФОТО: заголовок, url, описание в базы xlsx и txt
@app.route("/photo", methods=["POST", "GET"])
def photo():
    if request.method == "POST":
        title = request.form.get("title")
        urls = request.form.get("url")
        descriptions = request.form.get("description")
        # Запись в txt:
        photo_file = open("photo.txt", "a+", encoding="utf-8")
        photo_file.write(str(urls) + " " + str(title) + "\n")
        images = [[row[len(row.split()[0]):], row.split()[0]] for row in photo_file]
        photo_file.close()
        # Запись в xlsx лист Фото:
        excel = load_workbook("mybase.xlsx")
        page = excel["Фото"]
        page.append([title, urls, descriptions])
        excel.save("mybase.xlsx")
        return render_template("susses.html")
    
    # Вариант 1: Отображение через базу xlsx (более оптимальный)
    excel = load_workbook("mybase.xlsx")
    page = excel["Фото"]
    images = []
    for row in page:
        titlee = row[0].value
        urll = row[1].value
        descriptionss = row[2].value
        lst_images = [row[0].value, row[1].value, row[2].value]
        images.append(lst_images)

    # Вариант 2: Отображение через базу txt (неоптимальный)
    # photo_file = open("photo.txt", "r", encoding="utf-8")
    # images = [[row[len(row.split()[0]):], row.split()[0]] for row in photo_file]
    # photo_file.close()
    return render_template("photo.html", images=images)

#  Страница подробное отображения ФОТО: заголовок, url, описание в базы xlsx и txt
@app.route("/page/<number>")
def page(number):
    excel = load_workbook("mybase.xlsx")
    page = excel["Фото"]
    lst_row = page[number]
    return render_template("page.html", lst = lst_row)
